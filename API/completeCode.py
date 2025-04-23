import requests
import openpyxl
from openpyxl.utils import get_column_letter
import time
import os
from datetime import datetime

class PageType:
    HomePage = "HomePage"
    SmallSimplePage = "SmallSimplePage"
    HomeBuyingPage = "HomeBuyingPage"
    AboutPage = "AboutPage"
    USASearchPageCommunity = "USASearchPageCommunity"
    USASearchPagePlan = "USASearchPagePlan"
    USASearchPageQMI = "USASearchPageQMI"
    USAMPCPage = "USAMPCPage"
    USACommunityPage = "USACommunityPage"
    CANCondoPage = "CANCondoPage"
    CANCondoPlanPage = "CANCondoPlanPage"
    USAPlanPage = "USAPlanPage"
    USAInventoryPage = "USAInventoryPage"

class TestDataReader:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(self.file_path)
        self.sheet = self.workbook[self.sheet_name]
        self.header = [cell.value for cell in self.sheet[1]]

    def get_records_num(self):
        return self.sheet.max_row - 1  # Exclude header row

    def get_data(self, row_num, column_name):
        col_index = self.header.index(column_name) + 1  # +1 for 1-based index
        return self.sheet.cell(row=row_num + 1, column=col_index).value #+1 for header

class Config:
    def __init__(self, api_key, api_url, file_name):
        self.api_key = api_key
        self.api_url = api_url
        self.file_name = file_name
        self.cached_test_data_readers = {}

    def get_run_time_property(self, property_name):
        if property_name == "APIKey":
            return self.api_key
        elif property_name == "APIUrl":
            return self.api_url
        elif property_name == "FileName":
            return self.file_name
        return None

    def log_comment(self, message):
        print(message)

    def wait_without_logging(self, seconds):
        time.sleep(seconds)

    def get_cached_test_data_reader_object(self, sheet_name):
        if sheet_name not in self.cached_test_data_readers:
            file_path = os.path.join(os.getcwd(), "Parameters", self.file_name)
            self.cached_test_data_readers[sheet_name] = TestDataReader(file_path, sheet_name)
        return self.cached_test_data_readers[sheet_name]

class APIHelper:
    def __init__(self, test_config):
        self.test_config = test_config

    def submit_api_and_get_response(self, url, page_type, file_name):
        api_key = self.test_config.get_run_time_property("APIKey")
        api_url = self.test_config.get_run_time_property("APIUrl")

        params = {
            "url": url,
            "location": "Dulles:Chrome.FIOS",
            "f": "json",
            "runs": 2,
            "fvonly": 1,
            "lighthouse": 1,
        }
        headers = {"X-WPT-API-KEY": api_key}

        response = requests.get(api_url, params=params, verify=False)
        response.raise_for_status()
        response_json = response.json()

        user_url = response_json["data"]["userUrl"]
        api_json_url = response_json["data"]["jsonUrl"]

        json_response = requests.get(api_json_url, params={"f": "json"}, verify=False)
        json_response.raise_for_status()
        json_response_data = json_response.json()

        while str(json_response_data["statusCode"]) != "200":
            self.test_config.wait_without_logging(30)
            json_response = requests.get(api_json_url, params={"f": "json"}, verify=False)
            json_response.raise_for_status()
            json_response_data = json_response.json()

        self.test_config.log_comment(str(json_response_data["statusCode"]))
        self.test_config.log_comment(f"WebPageTest API Run URL: {api_json_url}")
        self.test_config.log_comment(f"WebPageTest API User URL: {user_url}")

        self.submit_values_in_csv(api_json_url, user_url, json_response_data, "First Run", page_type, file_name)
        self.submit_values_in_csv(api_json_url, user_url, json_response_data, "Second Run", page_type, file_name)

    def check_balance(self):
        api_key = self.test_config.get_run_time_property("APIKey")
        api_url = "https://webpagetest.org/testBalance.php"

        response = requests.get(api_url, headers={"X-WPT-API-KEY": api_key}, verify=False)
        response.raise_for_status()
        print(response.json())

    def submit_values_in_csv(self, api_url, user_url, response_json, run, page_type, file_name):

        run_num = 1 if run == "First Run" else 2

        load_time = response_json["data"]["runs"][str(run_num)]["firstView"]["loadTime"]
        ttfb = response_json["data"]["runs"][str(run_num)]["firstView"]["TTFB"]
        start_render = response_json["data"]["runs"][str(run_num)]["firstView"]["render"]
        speed_index_time = response_json["data"]["runs"][str(run_num)]["firstView"]["SpeedIndex"]
        document_requests_count = response_json["data"]["runs"][str(run_num)]["firstView"]["requestsDoc"]
        document_bytes_in = response_json["data"]["runs"][str(run_num)]["firstView"]["bytesInDoc"]
        document_time = response_json["data"]["runs"][str(run_num)]["firstView"]["docTime"]
        fully_loaded_requests_count = response_json["data"]["runs"][str(run_num)]["firstView"]["requestsFull"]
        full_bytes_in = response_json["data"]["runs"][str(run_num)]["firstView"]["bytesIn"]
        fully_loaded_time = response_json["data"]["runs"][str(run_num)]["firstView"]["fullyLoaded"]

        sheet_num = {
            PageType.HomePage: 0,
            PageType.SmallSimplePage: 1,
            PageType.HomeBuyingPage: 2,
            PageType.AboutPage: 3,
            PageType.USASearchPageCommunity: 4,
            PageType.USASearchPagePlan: 5,
            PageType.USASearchPageQMI: 6,
            PageType.USAMPCPage: 7,
            PageType.USACommunityPage: 8,
            PageType.CANCondoPage: 9,
            PageType.CANCondoPlanPage: 10,
            PageType.USAPlanPage: 11,
            PageType.USAInventoryPage: 12,
        }[page_type]

        file_path = os.path.join(os.getcwd(), "Parameters", file_name)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.worksheets[sheet_num]

        row_count = sheet.max_row
        row = sheet.insert_rows(row_count+1)

        data = [
            datetime.now().strftime("%m.%d.%y"),
            float(load_time) / 1000,
            float(ttfb) / 1000,
            float(start_render) / 1000,
            speed_index_time,
            float(document_time) / 1000,
            document_requests_count,
            f"{int(document_bytes_in) / 1024:,} KB",
            float(fully_loaded_time) / 1000,
            fully_loaded_requests_count,
            f"{int(full_bytes_in) / 1024:,} KB",
            run,
            user_url,
        ]

        for col_num, value in enumerate(data, start=1):
            sheet.cell(row=row_count+1, column=col_num, value=value)

        workbook.save(file_path)
        self.test_config.log_comment("File updated!!")

    def get_cell_values(self):
        file_name = self.test_config.get_run_time_property("FileName")
        file_path = os.path.join(os.getcwd(), "Parameters", file_name)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.worksheets[2]

        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 's':
                    print(f"[{cell.row}, {cell.column}] = STRING; Value = {cell.value}")
                elif cell.data_type == 'n':
                    print(f"[{cell.row}, {cell.column}] = NUMERIC; Value = {cell.value}")
                elif cell.data_type == 'b':
                    print(f"[{cell.row}, {cell.column}] = BOOLEAN; Value = {cell.value}")
                elif cell.data_type == 'e':
                    print(f"[{cell.row}, {cell.column}] = BLANK CELL")

    def read_values_and_extract_difference(self, page_type):
        sheet_name = {
            PageType.HomePage: "HomePage",
            PageType.SmallSimplePage: "SmallSimplePage",
            PageType.HomeBuyingPage: "HomeBuyingPage",
            PageType.AboutPage: "AboutPage",
            PageType.USASearchPageCommunity: "USASearchPageCommunity",
            PageType.USASearchPagePlan: "USASearchPagePlan",
            PageType.USASearchPageQMI: "USASearchPageQMI",
            PageType.USAMPCPage: "USAMPCPage",
            PageType.USACommunityPage: "USACommunityPage",
            PageType.CANCondoPage: "CANCondoPage",
            PageType.CANCondoPlanPage: "CANCondoPlanPage",
            PageType.USAPlanPage: "USAPlanPage",
            PageType.USAInventoryPage: "USAInventoryPage"
        }[page_type]

        heading_label = {
            PageType.HomePage: "Home Page",
            PageType.SmallSimplePage: "Small Simple Page",
            PageType.HomeBuyingPage: "HomeBuying Page",
            PageType.AboutPage: "About Page",
            PageType.USASearchPageCommunity: "USA Search Page Community",
            PageType.USASearchPagePlan: "USA Search Page Plan",
            PageType.USASearchPageQMI: "USA Search Page QMI",
            PageType.USAMPCPage: "USA MPC Page",
            PageType.USACommunityPage: "USA Community Page",
            PageType.CANCondoPage: "CAN Condo Page",
            PageType.CANCondoPlanPage: "CAN Condo Plan Page",
            PageType.USAPlanPage: "USA Plan Page",
            PageType.USAInventoryPage: "USA Inventory Page"
        }[page_type]

        category_reader = self.test_config.get_cached_test_data_reader_object(sheet_name)
        total_records = category_reader.get_records_num()
        self.test_config.log_comment(str(total_records))

        filename = f"PerformanceStats_{datetime.now().strftime('%dd-%mm-%yyyy')}.txt"
        file_path = os.path.join(os.getcwd(), "Parameters", filename)

        try:
            with open(file_path, "w") as fw:
                if total_records >= 5:
                    last_row_load_time = category_reader.get_data(total_records - 1, "Load Time")
                    last_row_ttfb = category_reader.get_data(total_records - 1, "First Byte")
                    last_row_fully_load_time = category_reader.get_data(total_records - 1, "FL Time")

                    third_last_row_load_time = category_reader.get_data(total_records - 3, "Load Time")
                    third_last_row_ttfb = category_reader.get_data(total_records - 3, "First Byte")
                    third_last_row_fully_load_time = category_reader.get_data(total_records - 3, "FL Time")

                    fw.write(f"{heading_label}\n")
                    fw.write(f"Load Time: {last_row_load_time} \n")

                    load_time_diff = float(last_row_load_time) - float(third_last_row_load_time)
                    fw.write(f"Load Time - Performance has {'declined' if load_time_diff > 0 else 'improved'} by {abs(load_time_diff):.2f} seconds \n")

                    load_to_fully_load_time_diff = float(last_row_fully_load_time) - float(third_last_row_fully_load_time)
                    fw.write(f"Time to Fully Loaded - Performance has {'declined' if load_to_fully_load_time_diff > 0 else 'improved'} by {abs(load_to_fully_load_time_diff):.2f} seconds \n")

                    ttfb_diff = float(last_row_ttfb) - float(third_last_row_ttfb)
                    fw.write(f"Time to First Byte - Performance has {'declined' if ttfb_diff > 0 else 'improved'} by {abs(ttfb_diff):.2f} seconds \n")
                else:
                    fw.write(f"{heading_label}\n")
                    fw.write(f"Not having enough records to compare... skipping comparison for {heading_label}\n")
        except Exception as e:
            print(f"An error occurred: {e}")

#Example usage
file_name = "example.xlsx" #replace with your file name.
api_key = "YOUR_API_KEY"
api_url = "https://webpagetest.org/runtest.php"

test_config = Config(api_key, api_url, file_name)
api_helper = APIHelper(test_config)

api_helper.submit_api_and_get_response("https://www.example.com", PageType.HomePage, file_name)
api_helper.get_cell_values()
api_helper.read_values_and_extract_difference(PageType.HomePage)